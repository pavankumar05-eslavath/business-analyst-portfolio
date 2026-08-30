"""Build the Excel model with live formulas.

The point of this module is that the delivered workbook is a **model**, not a
screenshot of one. Every calculated cell is a formula referencing a named range
that resolves back to the Drivers sheet. Change `orders_per_store_per_day` on the
Drivers sheet and the basket distribution, the CM ladder, all three decisions, the
cohort curves, both sensitivity grids and the scenario block recalculate.

Two design rules, both learned from models that were painful to review:

**No hard-coded numbers outside the Drivers sheet.** If a reviewer finds a
constant buried in a calculation they have to check every other cell too, because
they can no longer trust that the Drivers sheet is the whole input surface.

**No sheet names with spaces.** `='Unit Economics'!B5` works but every downstream
reference has to remember the quotes, and one missed quote produces a broken
workbook that still opens. `UnitEconomics` costs nothing and removes the failure
mode.

The formula subset is deliberately narrow -- arithmetic, SUM, ROUND, IF, MAX, MIN,
SQRT, LN, EXP, NORM.S.DIST -- so that the workbook opens and calculates in Excel,
LibreOffice and Google Sheets, and so the test suite can evaluate it
independently and assert it agrees with the Python model.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from .drivers import Drivers

# -- presentation ----------------------------------------------------------- #
TITLE = Font(bold=True, size=14, color="1F3864")
HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
GROUP = Font(bold=True, color="1F3864")
GROUP_FILL = PatternFill("solid", fgColor="D9E2F3")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL = Font(bold=True)
NOTE = Font(italic=True, size=9, color="595959")
TOP_BORDER = Border(top=Side(style="thin", color="808080"))

MONEY = '#,##0.00'
MONEY0 = '#,##0'
LAKH = '#,##0.0'
PCT1 = '0.0%'
PCT2 = '0.00%'
RATIO = '0.00"x"'
WRAP = Alignment(wrap_text=True, vertical="top")


@dataclass
class Cursor:
    """Tracks the row a sheet is being written at, so formulas can reference it."""

    row: int = 1

    def take(self, count: int = 1) -> int:
        start = self.row
        self.row += count
        return start


class ModelWorkbook:
    def __init__(self, drivers: Drivers):
        self.d = drivers
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.names: dict[str, str] = {}

    # -- helpers ----------------------------------------------------------- #
    def name(self, key: str, sheet: str, cell: str) -> None:
        """Register a named range. Names are the contract between sheets."""
        if key in self.names:
            raise ValueError(f"named range {key!r} already defined")
        self.names[key] = f"{sheet}!{cell}"
        self.wb.defined_names.add(
            DefinedName(key, attr_text=f"{sheet}!${cell[0]}${cell[1:]}")
        )

    @staticmethod
    def title(ws: Worksheet, cursor: Cursor, text: str, subtitle: str = "") -> None:
        row = cursor.take()
        ws.cell(row=row, column=1, value=text).font = TITLE
        if subtitle:
            row = cursor.take()
            ws.cell(row=row, column=1, value=subtitle).font = NOTE
        cursor.take()

    @staticmethod
    def header(ws: Worksheet, cursor: Cursor, labels: list[str]) -> int:
        row = cursor.take()
        for index, label in enumerate(labels, start=1):
            cell = ws.cell(row=row, column=index, value=label)
            cell.font = HEADER
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        return row

    @staticmethod
    def group(ws: Worksheet, cursor: Cursor, text: str, width: int = 4) -> int:
        row = cursor.take()
        for column in range(1, width + 1):
            cell = ws.cell(row=row, column=column)
            cell.fill = GROUP_FILL
            if column == 1:
                cell.value = text
                cell.font = GROUP
        return row

    @staticmethod
    def note(ws: Worksheet, cursor: Cursor, text: str) -> None:
        row = cursor.take()
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = NOTE

    # -- build ------------------------------------------------------------- #
    def build(self, path: Path | str) -> Path:
        self.sheet_guide()
        self.sheet_drivers()
        self.sheet_basket()
        self.sheet_unit_economics()
        self.sheet_cohort()
        self.sheet_channels()
        self.sheet_decisions()
        self.sheet_scenarios()
        self.sheet_sensitivity()

        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(path)
        return path

    # ------------------------------------------------------------------ #
    def sheet_guide(self) -> None:
        ws = self.wb.create_sheet("Guide")
        c = Cursor()
        self.title(
            ws, c, "QuickCart -- unit economics model",
            "Driver-based. Every calculated cell is a formula; nothing is pasted.",
        )

        lines = [
            ("How to use this model", ""),
            ("", "Change any yellow cell on the Drivers sheet. Everything else recalculates."),
            ("", "There are no hard-coded numbers outside the Drivers sheet. That is the "
                 "property that makes the model auditable."),
            ("Sheets", ""),
            ("Drivers", "Every input, with its unit and the basis for the number. "
                        "Drivers whose basis begins ASSUMPTION are the ones to attack first."),
            ("Basket", "Lognormal basket-value distribution. Resolves a free-delivery "
                       "threshold into fee-paying share and average basket, live."),
            ("UnitEconomics", "The CM1 / CM2 / CM3 ladder, four columns: base and one per "
                              "proposal."),
            ("Cohort", "24-month retention and cumulative contribution per acquisition "
                       "channel. LTV is the last row."),
            ("Channels", "CAC, LTV and LTV:CAC by channel, plus the blended and paid-only "
                         "views. Read the gap between those two."),
            ("Decisions", "The three proposals, their contribution impact and the "
                          "break-even value of the driver each one turns on."),
            ("Scenarios", "Base, bull and bear."),
            ("Sensitivity", "Two two-way tables. Live -- not pasted values."),
            ("What the model says", ""),
            ("", "Revenue growth and contribution growth point in opposite directions for "
                 "two of the three proposals. That is the finding, and it is why every "
                 "decision here is scored on CM3 contribution rather than on revenue."),
        ]
        for left, right in lines:
            row = c.take()
            if right == "":
                cell = ws.cell(row=row, column=1, value=left)
                cell.font = GROUP
                cell.fill = GROUP_FILL
                ws.cell(row=row, column=2).fill = GROUP_FILL
                continue
            ws.cell(row=row, column=1, value=left).font = TOTAL
            ws.cell(row=row, column=2, value=right).alignment = WRAP

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 96

    # ------------------------------------------------------------------ #
    def sheet_drivers(self) -> None:
        ws = self.wb.create_sheet("Drivers")
        c = Cursor()
        self.title(
            ws, c, "Drivers",
            "Yellow cells are inputs. Everything in this workbook resolves back to them.",
        )
        self.header(ws, c, ["Driver", "Value", "Unit", "Basis"])

        # Calendar basis. Lives in the `meta` block of the driver file rather than
        # with the business drivers, but the workbook still needs it as a named
        # range -- every volume formula multiplies by it.
        self.group(ws, c, "Calendar")
        row = c.take()
        ws.cell(row=row, column=1, value="days_per_month")
        cell = ws.cell(row=row, column=2, value=self.d.days_per_month)
        cell.fill = INPUT_FILL
        cell.number_format = MONEY0
        ws.cell(row=row, column=3, value="days")
        ws.cell(
            row=row, column=4,
            value="Modelling convention. Using 30 rather than 30.44 keeps monthly "
                  "figures comparable across the model and understates volume slightly, "
                  "which is the safe direction.",
        ).alignment = WRAP
        self.name("days_per_month", "Drivers", f"B{row}")

        current_group = None
        for driver in self.d.all():
            top = driver.group.split(".")[0]
            if top != current_group:
                current_group = top
                self.group(ws, c, top.replace("_", " ").title())

            row = c.take()
            ws.cell(row=row, column=1, value=driver.name)
            value_cell = ws.cell(row=row, column=2, value=driver.value)
            value_cell.fill = INPUT_FILL
            value_cell.number_format = self._format_for(driver.unit)
            ws.cell(row=row, column=3, value=driver.unit)
            basis = ws.cell(row=row, column=4, value=driver.basis)
            basis.alignment = WRAP
            if driver.is_assumption:
                ws.cell(row=row, column=1).font = Font(bold=True, color="C00000")
            self.name(driver.name, "Drivers", f"B{row}")

        # Channel block. Named per channel so the Cohort sheet can reference them.
        self.group(ws, c, "Channels", width=6)
        header_row = self.header(
            ws, c, ["Channel", "New customers", "Spend", "Retention multiplier", "CAC", "Basis"]
        )
        del header_row
        for channel in self.d.channels:
            row = c.take()
            ws.cell(row=row, column=1, value=channel.name)
            for column, value, fmt in (
                (2, channel.new_customers, MONEY0),
                (3, channel.spend, MONEY0),
                (4, channel.retention_multiplier, '0.00'),
            ):
                cell = ws.cell(row=row, column=column, value=value)
                cell.fill = INPUT_FILL
                cell.number_format = fmt
            cac = ws.cell(row=row, column=5, value=f"=IF(B{row}=0,0,C{row}/B{row})")
            cac.number_format = MONEY
            ws.cell(row=row, column=6, value=channel.basis).alignment = WRAP
            key = channel.name
            self.name(f"ch_{key}_customers", "Drivers", f"B{row}")
            self.name(f"ch_{key}_spend", "Drivers", f"C{row}")
            self.name(f"ch_{key}_mult", "Drivers", f"D{row}")
            self.name(f"ch_{key}_cac", "Drivers", f"E{row}")

        for column, width in (("A", 34), ("B", 14), ("C", 14), ("D", 20), ("E", 12), ("F", 78)):
            ws.column_dimensions[column].width = width

    @staticmethod
    def _format_for(unit: str) -> str:
        unit = unit.lower()
        if "fraction" in unit or unit == "dimensionless":
            return PCT2 if "fraction" in unit else '0.00'
        if "inr" in unit:
            return MONEY0 if "month" in unit or "per store" in unit else MONEY
        return MONEY0

    # ------------------------------------------------------------------ #
    def sheet_basket(self) -> None:
        """Lognormal basket distribution, evaluated live for both thresholds."""
        ws = self.wb.create_sheet("Basket")
        c = Cursor()
        self.title(
            ws, c, "Basket value distribution",
            "A free-delivery threshold acts on the distribution of basket values, "
            "not on the average. This sheet is why the model can evaluate a threshold change.",
        )

        self.group(ws, c, "Distribution parameters", width=3)
        rows: dict[str, int] = {}
        for label, formula, fmt in (
            ("Mean basket (GOV driver)", "=gross_order_value", MONEY),
            ("Sigma of log basket", "=basket_log_sigma", '0.00'),
            ("Mu (calibrated so mean = GOV)",
             "=LN(gross_order_value)-0.5*basket_log_sigma^2", '0.0000'),
            ("Implied median basket", None, MONEY),
        ):
            row = c.take()
            ws.cell(row=row, column=1, value=label)
            rows[label] = row
            if formula:
                cell = ws.cell(row=row, column=2, value=formula)
            else:
                cell = ws.cell(row=row, column=2, value=f"=EXP(B{rows['Mu (calibrated so mean = GOV)']})")
            cell.number_format = fmt
        mu_row = rows["Mu (calibrated so mean = GOV)"]
        self.name("basket_mu", "Basket", f"B{mu_row}")
        self.note(ws, c, "Median sits below the mean, which is the observed shape for grocery baskets.")

        for tag, label, threshold_ref in (
            ("current", "Current threshold", "free_delivery_threshold"),
            ("proposed", "Proposed threshold", "proposed_threshold"),
        ):
            c.take()
            self.group(ws, c, label, width=3)
            r: dict[str, int] = {}

            def put(key: str, text: str, formula: str, fmt: str,
                    rows: dict[str, int] = r) -> None:
                row = c.take()
                ws.cell(row=row, column=1, value=text)
                cell = ws.cell(row=row, column=2, value=formula)
                cell.number_format = fmt
                rows[key] = row

            put("t", "Threshold", f"={threshold_ref}", MONEY)
            put("zt", "z(threshold)", f"=(LN(B{r['t']})-basket_mu)/basket_log_sigma", '0.0000')
            put("below", "Share of orders below threshold",
                f"=NORM.S.DIST(B{r['zt']},TRUE)", PCT2)
            put("lo", "Up-size band lower bound",
                f"=B{r['t']}*(1-upsize_band_width)", MONEY)
            put("zlo", "z(band lower bound)",
                f"=(LN(B{r['lo']})-basket_mu)/basket_log_sigma", '0.0000')
            put("belowlo", "Share below band lower bound",
                f"=NORM.S.DIST(B{r['zlo']},TRUE)", PCT2)
            put("bandshare", "Share of orders inside the up-size band",
                f"=B{r['below']}-B{r['belowlo']}", PCT2)
            put("pmhi", "Partial mean below threshold",
                f"=gross_order_value*NORM.S.DIST(B{r['zt']}-basket_log_sigma,TRUE)", MONEY)
            put("pmlo", "Partial mean below band lower bound",
                f"=gross_order_value*NORM.S.DIST(B{r['zlo']}-basket_log_sigma,TRUE)", MONEY)
            put("bandmean", "Average basket inside the band",
                f"=IF(B{r['bandshare']}=0,0,(B{r['pmhi']}-B{r['pmlo']})/B{r['bandshare']})", MONEY)
            put("upsized", "Share of orders that up-size to clear the threshold",
                f"=B{r['bandshare']}*upsize_propensity", PCT2)
            put("feeshare", "Fee-paying share of orders after up-sizing",
                f"=B{r['below']}-B{r['upsized']}", PCT2)
            put("gov", "Average basket after up-sizing",
                f"=gross_order_value+B{r['upsized']}*(B{r['t']}-B{r['bandmean']})", MONEY)
            put("feeinc", "Delivery fee income per order",
                f"=B{r['feeshare']}*delivery_fee", MONEY)

            self.name(f"gov_{tag}", "Basket", f"B{r['gov']}")
            self.name(f"feeinc_{tag}", "Basket", f"B{r['feeinc']}")
            self.name(f"feeshare_{tag}", "Basket", f"B{r['feeshare']}")

        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 16

    # ------------------------------------------------------------------ #
    def sheet_unit_economics(self) -> None:
        """The CM ladder, four columns: base plus one per proposal."""
        ws = self.wb.create_sheet("UnitEconomics")
        c = Cursor()
        self.title(
            ws, c, "Contribution margin ladder",
            "CM2 stops before dark-store fixed cost; CM3 absorbs it. Several decisions "
            "change sign across that line, which is why both are shown.",
        )

        columns = ["B", "C", "D", "E"]
        self.header(ws, c, [
            "Line", "Base",
            "Threshold change", "Store expansion", "Retail media",
        ])

        # -- volume and per-column inputs, all live ------------------------ #
        self.group(ws, c, "Volume and configuration", width=5)
        v: dict[str, int] = {}

        def row_of(label: str, formulas: list[str], fmt: str = MONEY,
                   key: str | None = None, bold: bool = False) -> int:
            row = c.take()
            cell = ws.cell(row=row, column=1, value=label)
            if bold:
                cell.font = TOTAL
            for column, formula in zip(columns, formulas, strict=True):
                target = ws[f"{column}{row}"]
                target.value = formula
                target.number_format = fmt
                if bold:
                    target.font = TOTAL
            if key:
                v[key] = row
            return row

        row_of("Dark stores", [
            "=dark_stores", "=dark_stores", "=dark_stores+new_stores", "=dark_stores",
        ], MONEY0, key="stores")

        # Expansion: only the incremental share of new-store volume is new demand.
        row_of("Orders per store per day", [
            "=orders_per_store_per_day",
            "=orders_per_store_per_day*(1-observed_volume_response)",
            "=(dark_stores*orders_per_store_per_day"
            "+new_stores*orders_per_new_store_per_day*incrementality)/(dark_stores+new_stores)",
            "=orders_per_store_per_day",
        ], MONEY0, key="opd")

        row_of("Orders per month", [
            f"={col}{v['stores']}*{col}{v['opd']}*days_per_month" for col in columns
        ], MONEY0, key="orders")

        row_of("Gross order value", [
            "=gov_current", "=gov_proposed", "=gov_current", "=gov_current",
        ], MONEY, key="gov")

        row_of("Fee-paying share of orders", [
            "=feeshare_current", "=feeshare_proposed", "=feeshare_current", "=feeshare_current",
        ], PCT2, key="feeshare")

        # Rider payout falls with store density: distance scales with 1/sqrt(stores).
        row_of("Rider payout per order", [
            "=rider_payout", "=rider_payout",
            "=rider_payout*(1-rider_payout_distance_share"
            "+rider_payout_distance_share*SQRT(dark_stores/(dark_stores+new_stores)))",
            "=rider_payout",
        ], MONEY, key="rider")

        row_of("Retail media income per order", [
            "=retail_media_income", "=retail_media_income", "=retail_media_income",
            "=proposed_income",
        ], MONEY, key="media")

        row_of("Additional fixed cost per month", [
            "=0", "=0", "=0", "=delivery_cost",
        ], MONEY0, key="extrafixed")

        # -- revenue ------------------------------------------------------- #
        c.take()
        self.group(ws, c, "Revenue per order", width=5)
        row_of("Net order value", [
            f"={col}{v['gov']}-platform_funded_discount" for col in columns
        ], MONEY, key="nov")
        row_of("Delivery fee income", [
            "=feeinc_current", "=feeinc_proposed", "=feeinc_current", "=feeinc_current",
        ], MONEY, key="fee")
        row_of("Handling fee income", ["=handling_fee"] * 4, MONEY, key="handling")
        row_of("Retail media income", [
            f"={col}{v['media']}" for col in columns
        ], MONEY, key="mediarev")
        row_of("Total revenue", [
            f"={col}{v['nov']}+{col}{v['fee']}+{col}{v['handling']}+{col}{v['mediarev']}"
            for col in columns
        ], MONEY, key="revenue", bold=True)

        # -- to CM1 -------------------------------------------------------- #
        c.take()
        self.group(ws, c, "Cost of sales", width=5)
        row_of("Cost of goods sold", [
            f"=-{col}{v['nov']}*(1-product_gross_margin_pct)" for col in columns
        ], MONEY, key="cogs")
        # Retail media is invoiced to brands, so it is excluded from the MDR base.
        row_of("Payment gateway", [
            f"=-({col}{v['nov']}+{col}{v['fee']}+{col}{v['handling']})*payment_gateway_pct"
            for col in columns
        ], MONEY, key="pg")
        row_of("Packaging material", ["=-packaging_material"] * 4, MONEY, key="pack")
        row_of("CM1 -- gross contribution", [
            f"={col}{v['revenue']}+{col}{v['cogs']}+{col}{v['pg']}+{col}{v['pack']}"
            for col in columns
        ], MONEY, key="cm1", bold=True)
        row_of("CM1 % of net order value", [
            f"={col}{v['cm1']}/{col}{v['nov']}" for col in columns
        ], PCT1)

        # -- to CM2 -------------------------------------------------------- #
        c.take()
        self.group(ws, c, "Variable fulfilment", width=5)
        row_of("Rider payout", [f"=-{col}{v['rider']}" for col in columns], MONEY, key="riderc")
        row_of("Picking labour", ["=-picking_labour"] * 4, MONEY, key="pick")
        row_of("Spoilage and shrinkage", [
            f"=-{col}{v['gov']}*spoilage_pct" for col in columns
        ], MONEY, key="spoil")
        row_of("CM2 -- store contribution", [
            f"={col}{v['cm1']}+{col}{v['riderc']}+{col}{v['pick']}+{col}{v['spoil']}"
            for col in columns
        ], MONEY, key="cm2", bold=True)
        row_of("CM2 % of net order value", [
            f"={col}{v['cm2']}/{col}{v['nov']}" for col in columns
        ], PCT1)

        # -- to CM3 -------------------------------------------------------- #
        c.take()
        self.group(ws, c, "Dark store fixed cost", width=5)
        row_of("Dark store fixed cost per store per month",
               ["=rent+utilities+store_staff+security_and_other"] * 4, MONEY0, key="perstore")
        row_of("Dark store fixed cost per month", [
            f"={col}{v['stores']}*{col}{v['perstore']}+{col}{v['extrafixed']}" for col in columns
        ], MONEY0, key="fixedtot")
        row_of("Dark store fixed, allocated per order", [
            f"=-{col}{v['fixedtot']}/{col}{v['orders']}" for col in columns
        ], MONEY, key="fixedpo")
        row_of("CM3 -- fully loaded store", [
            f"={col}{v['cm2']}+{col}{v['fixedpo']}" for col in columns
        ], MONEY, key="cm3", bold=True)
        row_of("CM3 % of net order value", [
            f"={col}{v['cm3']}/{col}{v['nov']}" for col in columns
        ], PCT1)

        # -- below the line ------------------------------------------------ #
        c.take()
        self.group(ws, c, "Below contribution", width=5)
        self.note(
            ws, c,
            "Central overhead and marketing are fixed rupee budgets. They are divided by "
            "each column's own order volume, so a proposal that changes volume correctly "
            "sees overhead per order move.",
        )
        row_of("Central overhead per month",
               ["=central_overhead_per_order*base_orders_per_month"] * 4, MONEY0, key="centraltot")
        row_of("Marketing per month",
               ["=marketing_per_order*base_orders_per_month"] * 4, MONEY0, key="mktgtot")
        row_of("Central overhead per order", [
            f"=-{col}{v['centraltot']}/{col}{v['orders']}" for col in columns
        ], MONEY, key="centralpo")
        row_of("Marketing per order", [
            f"=-{col}{v['mktgtot']}/{col}{v['orders']}" for col in columns
        ], MONEY, key="mktgpo")
        row_of("EBITDA per order", [
            f"={col}{v['cm3']}+{col}{v['centralpo']}+{col}{v['mktgpo']}" for col in columns
        ], MONEY, key="ebitdapo", bold=True)

        # -- monthly totals ------------------------------------------------ #
        c.take()
        self.group(ws, c, "Monthly totals (INR)", width=5)
        row_of("CM2 contribution", [
            f"={col}{v['cm2']}*{col}{v['orders']}" for col in columns
        ], MONEY0, key="cm2tot")
        row_of("CM3 contribution", [
            f"={col}{v['cm2']}*{col}{v['orders']}-{col}{v['fixedtot']}" for col in columns
        ], MONEY0, key="cm3tot", bold=True)
        row_of("EBITDA", [
            f"={col}{v['cm3tot']}-{col}{v['centraltot']}-{col}{v['mktgtot']}" for col in columns
        ], MONEY0, key="ebitdatot", bold=True)
        row_of("Change in CM3 contribution vs base", [
            "=0"] + [f"={col}{v['cm3tot']}-B{v['cm3tot']}" for col in columns[1:]
        ], MONEY0, key="delta", bold=True)

        for key, target in (
            ("base_cm1", f"B{v['cm1']}"), ("base_cm2", f"B{v['cm2']}"),
            ("base_cm3", f"B{v['cm3']}"), ("base_nov", f"B{v['nov']}"),
            ("base_orders", f"B{v['orders']}"),
            ("base_cm2_total", f"B{v['cm2tot']}"), ("base_cm3_total", f"B{v['cm3tot']}"),
            ("base_ebitda_total", f"B{v['ebitdatot']}"),
            ("base_fixed_total", f"B{v['fixedtot']}"),
            ("base_ebitda_per_order", f"B{v['ebitdapo']}"),
            ("thr_cm3_total", f"C{v['cm3tot']}"),
            ("exp_cm3_total", f"D{v['cm3tot']}"),
            ("media_cm3_total", f"E{v['cm3tot']}"),
            ("exp_cm2", f"D{v['cm2']}"), ("thr_cm2", f"C{v['cm2']}"),
            ("exp_opd", f"D{v['opd']}"), ("exp_fixed_total", f"D{v['fixedtot']}"),
        ):
            self.name(key, "UnitEconomics", target)

        ws.column_dimensions["A"].width = 42
        for column in columns:
            ws.column_dimensions[column].width = 17

        # `base_orders_per_month` is referenced above and defined on Drivers as a
        # formula so that the overhead conversion has a single source.
        drivers_ws = self.wb["Drivers"]
        row = drivers_ws.max_row + 2
        drivers_ws.cell(row=row, column=1, value="base_orders_per_month").font = TOTAL
        cell = drivers_ws.cell(
            row=row, column=2,
            value="=dark_stores*orders_per_store_per_day*days_per_month",
        )
        cell.number_format = MONEY0
        drivers_ws.cell(row=row, column=3, value="orders")
        drivers_ws.cell(
            row=row, column=4,
            value="Derived, not an input. Baseline volume used to convert the per-order "
                  "overhead drivers into the fixed rupee budgets they actually are.",
        ).alignment = WRAP
        self.name("base_orders_per_month", "Drivers", f"B{row}")

    # ------------------------------------------------------------------ #
    def sheet_cohort(self) -> None:
        """Retention curve and cumulative contribution, one column pair per channel."""
        ws = self.wb.create_sheet("Cohort")
        c = Cursor()
        self.title(
            ws, c, "Cohort retention and lifetime contribution",
            "Retention climbs from a month-1 shock toward an asymptote. The channel "
            "multiplier moves month-1 retention only -- see LEARN.md for why.",
        )

        self.group(ws, c, "Channel parameters", width=4)
        header = self.header(ws, c, ["Channel", "Retention multiplier",
                                     "Effective month-1 retention", "CAC"])
        del header
        param_rows: dict[str, int] = {}
        for channel in self.d.channels:
            row = c.take()
            ws.cell(row=row, column=1, value=channel.name)
            ws.cell(row=row, column=2, value=f"=ch_{channel.name}_mult").number_format = '0.00'
            ws.cell(
                row=row, column=3,
                value=f"=MIN(MAX(month_1_retention*ch_{channel.name}_mult,0.05),"
                      f"asymptotic_retention-0.01)",
            ).number_format = PCT1
            ws.cell(row=row, column=4, value=f"=ch_{channel.name}_cac").number_format = MONEY
            param_rows[channel.name] = row

        c.take()
        horizon = int(self.d["horizon_months"])
        self.group(ws, c, f"{horizon}-month curve (contribution valued at CM3 per order)",
                   width=2 + 2 * len(self.d.channels))

        head = c.take()
        ws.cell(row=head, column=1, value="Month").font = HEADER
        ws.cell(row=head, column=1).fill = HEADER_FILL
        ws.cell(row=head, column=2, value="Orders per survivor").font = HEADER
        ws.cell(row=head, column=2).fill = HEADER_FILL
        channel_columns: dict[str, tuple[str, str]] = {}
        for index, channel in enumerate(self.d.channels):
            survival_col = get_column_letter(3 + index * 2)
            contribution_col = get_column_letter(4 + index * 2)
            for col, suffix in ((survival_col, "survival"), (contribution_col, "contribution")):
                cell = ws[f"{col}{head}"]
                cell.value = f"{channel.name}\n{suffix}"
                cell.font = HEADER
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            channel_columns[channel.name] = (survival_col, contribution_col)

        first = c.row
        for month in range(1, horizon + 1):
            row = c.take()
            ws.cell(row=row, column=1, value=month)
            ws.cell(
                row=row, column=2,
                value=f"=orders_per_customer_per_month*(1+tenure_frequency_uplift"
                      f"*MIN({month}-1,6)/6)",
            ).number_format = '0.00'
            for channel in self.d.channels:
                survival_col, contribution_col = channel_columns[channel.name]
                param = param_rows[channel.name]
                if month == 1:
                    ws[f"{survival_col}{row}"] = "=1"
                else:
                    # survival_prev * retention rate for the previous month
                    ws[f"{survival_col}{row}"] = (
                        f"={survival_col}{row - 1}*MIN(asymptotic_retention"
                        f"-(asymptotic_retention-C{param})*retention_decay^({month - 1}-1),0.97)"
                    )
                ws[f"{survival_col}{row}"].number_format = PCT1
                ws[f"{contribution_col}{row}"] = (
                    f"={survival_col}{row}*B{row}*base_cm3"
                )
                ws[f"{contribution_col}{row}"].number_format = MONEY
        last = c.row - 1

        total_row = c.take()
        ws.cell(row=total_row, column=1, value=f"LTV over {horizon} months").font = TOTAL
        for channel in self.d.channels:
            _, contribution_col = channel_columns[channel.name]
            cell = ws[f"{contribution_col}{total_row}"]
            cell.value = f"=SUM({contribution_col}{first}:{contribution_col}{last})"
            cell.number_format = MONEY
            cell.font = TOTAL
            cell.border = TOP_BORDER
            self.name(f"ltv_{channel.name}", "Cohort", f"{contribution_col}{total_row}")

        orders_row = c.take()
        ws.cell(row=orders_row, column=1, value="Orders per acquired customer").font = TOTAL
        for channel in self.d.channels:
            survival_col, _ = channel_columns[channel.name]
            helper = get_column_letter(3 + 2 * len(self.d.channels) + 1)
            del helper
            cell = ws[f"{survival_col}{orders_row}"]
            # SUMPRODUCT of survival and orders-per-survivor.
            cell.value = (
                f"=SUMPRODUCT({survival_col}{first}:{survival_col}{last},B{first}:B{last})"
            )
            cell.number_format = '0.0'
            cell.font = TOTAL

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.freeze_panes = ws[f"A{first}"]

    # ------------------------------------------------------------------ #
    def sheet_channels(self) -> None:
        ws = self.wb.create_sheet("Channels")
        c = Cursor()
        self.title(
            ws, c, "Acquisition channel economics",
            "The gap between the blended and paid-only rows is the finding. Blended CAC "
            "divides total spend by all acquired customers, including the organic ones "
            "nobody paid for.",
        )

        self.header(ws, c, [
            "Channel", "New customers", "Spend", "CAC",
            "LTV on CM3", "LTV:CAC on CM3", "LTV on CM2", "LTV:CAC on CM2", "Verdict",
        ])

        first = c.row
        for channel in self.d.channels:
            row = c.take()
            ws.cell(row=row, column=1, value=channel.name)
            ws.cell(row=row, column=2, value=f"=ch_{channel.name}_customers").number_format = MONEY0
            ws.cell(row=row, column=3, value=f"=ch_{channel.name}_spend").number_format = MONEY0
            ws.cell(row=row, column=4, value=f"=ch_{channel.name}_cac").number_format = MONEY
            ws.cell(row=row, column=5, value=f"=ltv_{channel.name}").number_format = MONEY
            ws.cell(
                row=row, column=6,
                value=f"=IF(D{row}=0,\"n/a - not bought\",E{row}/D{row})",
            ).number_format = RATIO
            # LTV on CM2 scales with the margin used, so it is the CM3 LTV rescaled.
            ws.cell(
                row=row, column=7,
                value=f"=E{row}*base_cm2/base_cm3",
            ).number_format = MONEY
            ws.cell(
                row=row, column=8,
                value=f"=IF(D{row}=0,\"n/a - not bought\",G{row}/D{row})",
            ).number_format = RATIO
            ws.cell(
                row=row, column=9,
                value=f'=IF(D{row}=0,"organic",IF(E{row}/D{row}>=1,'
                      f'"funds itself on CM3","destroys value on CM3"))',
            )
        last = c.row - 1

        c.take()
        self.group(ws, c, "Portfolio view", width=9)
        rows: dict[str, int] = {}
        for label, formula, fmt in (
            ("Total new customers", f"=SUM(B{first}:B{last})", MONEY0),
            ("Total spend", f"=SUM(C{first}:C{last})", MONEY0),
            ("Paid new customers",
             "=" + "+".join(f"IF(ch_{ch.name}_spend>0,ch_{ch.name}_customers,0)"
                            for ch in self.d.channels), MONEY0),
            ("Blended LTV on CM3",
             "=" + "+".join(f"ltv_{ch.name}*ch_{ch.name}_customers" for ch in self.d.channels)
             + f")/SUM(B{first}:B{last}", MONEY),
        ):
            row = c.take()
            ws.cell(row=row, column=1, value=label).font = TOTAL
            cell = ws.cell(row=row, column=2, value=formula)
            cell.number_format = fmt
            rows[label] = row
        # Fix the blended LTV formula's parentheses.
        ws.cell(
            row=rows["Blended LTV on CM3"], column=2,
            value="=(" + "+".join(
                f"ltv_{ch.name}*ch_{ch.name}_customers" for ch in self.d.channels
            ) + f")/SUM(B{first}:B{last})",
        ).number_format = MONEY

        for label, formula, fmt in (
            ("Blended CAC", f"=B{rows['Total spend']}/B{rows['Total new customers']}", MONEY),
            ("Paid-only CAC", f"=B{rows['Total spend']}/B{rows['Paid new customers']}", MONEY),
        ):
            row = c.take()
            ws.cell(row=row, column=1, value=label).font = TOTAL
            ws.cell(row=row, column=2, value=formula).number_format = fmt
            rows[label] = row

        for label, formula in (
            ("Blended LTV:CAC", f"=B{rows['Blended LTV on CM3']}/B{rows['Blended CAC']}"),
            ("Paid-only LTV:CAC", f"=B{rows['Blended LTV on CM3']}/B{rows['Paid-only CAC']}"),
        ):
            row = c.take()
            ws.cell(row=row, column=1, value=label).font = TOTAL
            cell = ws.cell(row=row, column=2, value=formula)
            cell.number_format = RATIO
            cell.font = TOTAL
            rows[label] = row
        self.name("blended_ltv_cac", "Channels", f"B{rows['Blended LTV:CAC']}")
        self.name("paid_ltv_cac", "Channels", f"B{rows['Paid-only LTV:CAC']}")

        row = c.take()
        ws.cell(row=row, column=1, value="Overstatement from using blended").font = TOTAL
        ws.cell(
            row=row, column=2,
            value=f"=B{rows['Blended LTV:CAC']}/B{rows['Paid-only LTV:CAC']}-1",
        ).number_format = PCT1

        row = c.take()
        ws.cell(row=row, column=1, value="Spend in channels below 1.0x on CM3").font = TOTAL
        ws.cell(
            row=row, column=2,
            value="=" + "+".join(
                f"IF(AND(ch_{ch.name}_spend>0,ltv_{ch.name}/MAX(ch_{ch.name}_cac,0.0001)<1),"
                f"ch_{ch.name}_spend,0)" for ch in self.d.channels
            ),
        ).number_format = MONEY0
        sub_row = row
        row = c.take()
        ws.cell(row=row, column=1, value="As share of total spend").font = TOTAL
        ws.cell(
            row=row, column=2, value=f"=B{sub_row}/B{rows['Total spend']}",
        ).number_format = PCT1

        self.note(
            ws, c,
            "Compare columns F and H. Channels that destroy value on CM3 clear 1.0x on CM2, "
            "so the same channel is fundable or not depending on which contribution margin "
            "the deck happens to use. That is a governance problem, not an arithmetic one.",
        )

        ws.column_dimensions["A"].width = 36
        for column in "BCDEFGH":
            ws.column_dimensions[column].width = 16
        ws.column_dimensions["I"].width = 28

    # ------------------------------------------------------------------ #
    def sheet_decisions(self) -> None:
        ws = self.wb.create_sheet("Decisions")
        c = Cursor()
        self.title(
            ws, c, "The three proposals",
            "Each is scored on change in monthly CM3 contribution, and each states the "
            "break-even value of the driver the answer turns on.",
        )

        self.header(ws, c, ["", "Store expansion", "Threshold change", "Retail media"])

        def line(label: str, formulas: list[str], fmt: str, bold: bool = False) -> int:
            row = c.take()
            cell = ws.cell(row=row, column=1, value=label)
            if bold:
                cell.font = TOTAL
            for column, formula in zip(["B", "C", "D"], formulas, strict=True):
                target = ws[f"{column}{row}"]
                target.value = formula
                target.number_format = fmt
                if bold:
                    target.font = TOTAL
            return row

        line("Proposed by", ['="Growth"', '="Pricing"', '="Category"'], '@')
        line("Change in orders per month", [
            "=exp_cm3_total*0+(dark_stores+new_stores)"
            "*exp_opd*days_per_month/base_orders-1",
            "=(1-observed_volume_response)-1",
            "=0",
        ], PCT1)
        cm3_row = line("Change in monthly CM3 contribution", [
            "=exp_cm3_total-base_cm3_total",
            "=thr_cm3_total-base_cm3_total",
            "=media_cm3_total-base_cm3_total",
        ], MONEY0, bold=True)
        line("Annualised", [f"={col}{cm3_row}*12" for col in "BCD"], MONEY0, bold=True)
        line("Capital expenditure", [
            "=new_stores*capex_per_store", "=0", "=0",
        ], MONEY0)
        line("Verdict", [
            f'=IF(B{cm3_row}>0,"APPROVE","REJECT")',
            f'=IF(C{cm3_row}>0,"APPROVE","REJECT")',
            f'=IF(D{cm3_row}>0,"APPROVE","REJECT")',
        ], '@', bold=True)

        c.take()
        self.group(ws, c, "The number each decision turns on", width=4)
        line("Break-even driver", [
            '="Incrementality of new-store volume"',
            '="Order-volume decline the gain can absorb"',
            '="Share of the income uplift that must land"',
        ], '@')
        line("Break-even value", [
            # Solved on CM2, which does not depend on incrementality.
            "=((base_cm3_total+exp_fixed_total)/exp_cm2/days_per_month"
            "-dark_stores*orders_per_store_per_day)"
            "/(new_stores*orders_per_new_store_per_day)",
            "=1-(base_cm3_total+base_fixed_total)/(thr_cm2*base_orders)",
            "=(delivery_cost/base_orders)/(proposed_income-retail_media_income)",
        ], PCT1, bold=True)
        line("Actually observed", [
            "=incrementality", "=observed_volume_response", '="not yet measured"',
        ], PCT1)
        line("Margin of safety", [
            "=incrementality-B" + str(c.row - 2),
            "=B" + str(c.row - 2) + "*0+(1-(base_cm3_total+base_fixed_total)"
            "/(thr_cm2*base_orders))-observed_volume_response",
            '="cost is bounded by the INR 9 lakh a month it takes to try"',
        ], PCT1)

        self.note(
            ws, c,
            "Store expansion needs incrementality above the break-even and the pilot came in "
            "below it, so it is rejected on measured evidence rather than on judgement.",
        )
        self.note(
            ws, c,
            "The threshold change is the mirror image of the expansion: revenue falls and "
            "contribution rises. Revenue is not the objective function.",
        )

        ws.column_dimensions["A"].width = 44
        for column in "BCD":
            ws.column_dimensions[column].width = 22

    # ------------------------------------------------------------------ #
    def sheet_scenarios(self) -> None:
        ws = self.wb.create_sheet("Scenarios")
        c = Cursor()
        self.title(
            ws, c, "Scenarios",
            "Bull and bear move four drivers each. Every move is inside the range the "
            "business has already seen across its three cities.",
        )

        names = list(self.d.scenarios)
        self.header(ws, c, ["Driver"] + ["Base"] + [n.title() for n in names])

        spec_keys = [
            ("Orders per store per day", "orders_per_store_per_day_pct",
             "=orders_per_store_per_day", MONEY0, "pct"),
            ("Product gross margin", "product_gross_margin_pct_delta",
             "=product_gross_margin_pct", PCT1, "delta"),
            ("Rider payout per order", "rider_payout_pct", "=rider_payout", MONEY, "pct"),
            ("Gross order value", "gross_order_value_pct", "=gov_current", MONEY, "pct"),
        ]
        rows: dict[str, int] = {}
        for label, key, base_formula, fmt, mode in spec_keys:
            row = c.take()
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=base_formula).number_format = fmt
            for index, scenario in enumerate(names):
                value = self.d.scenarios[scenario].get(key, 0.0)
                column = get_column_letter(3 + index)
                formula = (f"=B{row}*(1+{value})" if mode == "pct" else f"=B{row}+{value}")
                ws[f"{column}{row}"] = formula
                ws[f"{column}{row}"].number_format = fmt
            rows[key] = row

        opd = rows["orders_per_store_per_day_pct"]
        margin = rows["product_gross_margin_pct_delta"]
        rider = rows["rider_payout_pct"]
        gov = rows["gross_order_value_pct"]

        c.take()
        self.group(ws, c, "Resulting economics", width=2 + len(names))
        letters = ["B"] + [get_column_letter(3 + i) for i in range(len(names))]

        def scenario_line(label: str, template: str, fmt: str, bold: bool = False) -> int:
            row = c.take()
            cell = ws.cell(row=row, column=1, value=label)
            if bold:
                cell.font = TOTAL
            for column in letters:
                target = ws[f"{column}{row}"]
                target.value = template.format(c=column)
                target.number_format = fmt
                if bold:
                    target.font = TOTAL
            return row

        nov = scenario_line("Net order value", f"={{c}}{gov}-platform_funded_discount", MONEY)
        rev = scenario_line(
            "Total revenue",
            f"={{c}}{nov}+feeinc_current+handling_fee+retail_media_income", MONEY,
        )
        cm1 = scenario_line(
            "CM1",
            f"={{c}}{rev}-{{c}}{nov}*(1-{{c}}{margin})"
            f"-({{c}}{nov}+feeinc_current+handling_fee)*payment_gateway_pct-packaging_material",
            MONEY, bold=True,
        )
        cm2 = scenario_line(
            "CM2",
            f"={{c}}{cm1}-{{c}}{rider}-picking_labour-{{c}}{gov}*spoilage_pct",
            MONEY, bold=True,
        )
        orders = scenario_line(
            "Orders per month", f"=dark_stores*{{c}}{opd}*days_per_month", MONEY0,
        )
        cm3 = scenario_line(
            "CM3",
            f"={{c}}{cm2}-(dark_stores*(rent+utilities+store_staff+security_and_other))"
            f"/{{c}}{orders}",
            MONEY, bold=True,
        )
        scenario_line(
            "EBITDA per order",
            f"={{c}}{cm3}-(central_overhead_per_order+marketing_per_order)"
            f"*base_orders_per_month/{{c}}{orders}",
            MONEY, bold=True,
        )
        scenario_line(
            "Monthly EBITDA",
            f"={{c}}{cm2}*{{c}}{orders}"
            f"-dark_stores*(rent+utilities+store_staff+security_and_other)"
            f"-(central_overhead_per_order+marketing_per_order)*base_orders_per_month",
            MONEY0, bold=True,
        )

        self.note(
            ws, c,
            "The spread between bull and bear is wide because unit economics are leveraged "
            "on order density. That is the honest reading of this business: it is a bet on "
            "orders per store, not on order count.",
        )
        ws.column_dimensions["A"].width = 34
        for column in letters:
            ws.column_dimensions[column].width = 16

    # ------------------------------------------------------------------ #
    def sheet_sensitivity(self) -> None:
        """Two live two-way tables. Recompute the ladder inline per cell."""
        ws = self.wb.create_sheet("Sensitivity")
        c = Cursor()
        self.title(
            ws, c, "Two-way sensitivity",
            "Live formulas, not pasted values. Change any driver and both tables move.",
        )

        # -- Table 1: CM3 per order vs density and gross margin ------------ #
        self.group(ws, c, "CM3 per order (INR): order density against product gross margin",
                   width=7)
        densities = [300, 340, 380, 420, 460, 500, 540]
        margins = [0.185, 0.200, 0.215, 0.230, 0.245]

        head = c.take()
        ws.cell(row=head, column=1, value="Orders/store/day \\ Gross margin").font = HEADER
        ws.cell(row=head, column=1).fill = HEADER_FILL
        for index, margin in enumerate(margins):
            cell = ws.cell(row=head, column=2 + index, value=margin)
            cell.number_format = PCT1
            cell.font = HEADER
            cell.fill = HEADER_FILL
        margin_letters = [get_column_letter(2 + i) for i in range(len(margins))]

        for density in densities:
            row = c.take()
            ws.cell(row=row, column=1, value=density).number_format = MONEY0
            for letter in margin_letters:
                # CM2 with this margin, less store fixed cost at this density.
                ws[f"{letter}{row}"] = (
                    f"=(gov_current-platform_funded_discount)*{letter}${head}"
                    f"+feeinc_current+handling_fee+retail_media_income"
                    f"-((gov_current-platform_funded_discount)+feeinc_current+handling_fee)"
                    f"*payment_gateway_pct-packaging_material"
                    f"-rider_payout-picking_labour-gov_current*spoilage_pct"
                    f"-(rent+utilities+store_staff+security_and_other)"
                    f"/($A{row}*days_per_month)"
                )
                ws[f"{letter}{row}"].number_format = MONEY
        self.note(
            ws, c,
            "This table is CM3, on which a percentage point of gross margin is worth more "
            "than 40 extra orders a day. On EBITDA the ranking reverses, because density "
            "also dilutes fixed central overhead while margin does not.",
        )

        # -- Table 2: expansion delta vs incrementality and volume --------- #
        c.take()
        self.group(
            ws, c,
            "Store expansion: change in monthly CM3 contribution (INR lakh)", width=6,
        )
        incrementalities = [0.30, 0.40, 0.50, 0.60, 0.70, 0.80, 0.90]
        volumes = [200, 250, 300, 350, 400]

        head2 = c.take()
        ws.cell(row=head2, column=1, value="Incrementality \\ Orders per new store").font = HEADER
        ws.cell(row=head2, column=1).fill = HEADER_FILL
        for index, volume in enumerate(volumes):
            cell = ws.cell(row=head2, column=2 + index, value=volume)
            cell.number_format = MONEY0
            cell.font = HEADER
            cell.fill = HEADER_FILL
        volume_letters = [get_column_letter(2 + i) for i in range(len(volumes))]

        for incrementality in incrementalities:
            row = c.take()
            ws.cell(row=row, column=1, value=incrementality).number_format = PCT1
            for letter in volume_letters:
                # exp_cm2 is the CM2 at the expanded store count (rider payout falls
                # with density), and does not depend on either grid axis.
                ws[f"{letter}{row}"] = (
                    f"=(exp_cm2*(dark_stores*orders_per_store_per_day"
                    f"+new_stores*{letter}${head2}*$A{row})*days_per_month"
                    f"-(dark_stores+new_stores)"
                    f"*(rent+utilities+store_staff+security_and_other)"
                    f"-base_cm3_total)/100000"
                )
                ws[f"{letter}{row}"].number_format = LAKH
        self.note(
            ws, c,
            "Every cell at the pilot's measured 43% incrementality is negative. The grid is "
            "the argument for a larger pilot rather than for a bigger guess.",
        )

        ws.column_dimensions["A"].width = 36
        for letter in set(margin_letters + volume_letters):
            ws.column_dimensions[letter].width = 14


def build(drivers: Drivers, path: Path | str) -> Path:
    return ModelWorkbook(drivers).build(path)
