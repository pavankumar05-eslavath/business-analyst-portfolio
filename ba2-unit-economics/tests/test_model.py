"""Tests for the unit economics model.

Four groups, in increasing order of how much they would embarrass me if they failed:

1. **Arithmetic identities.** The ladder adds up, the distribution integrates to the
   mean, the totals reconcile with the per-order figures. Cheap and boring.
2. **Modelling conventions.** Chiefly that fixed costs stay fixed in rupees when
   volume moves. This is the convention that decides the store-expansion verdict, so
   it gets an explicit test that fails if someone "simplifies" it.
3. **The headline findings.** Every number quoted in README.md, DECISION.md and
   LEARN.md is pinned here. If a driver changes and a claim in the prose stops being
   true, the build fails rather than the document quietly becoming wrong.
4. **The delivered workbook.** The xlsx is evaluated by an independent formula engine
   and asserted to agree with the Python model, cell by named cell, and to contain no
   error cells anywhere. A model that ships as a spreadsheet has to be verified as a
   spreadsheet -- verifying the Python that wrote it proves nothing about what a
   recruiter opens.
"""

from __future__ import annotations

import math
import warnings

import pytest
import yaml

from src import drivers as drivers_module
from src.decisions import (
    build_scenarios,
    cm3_sensitivity,
    compute_break_even,
    evaluate_all,
    evaluate_combined,
    expansion_sensitivity,
)
from src.model import (
    BasketDistribution,
    build_channel_portfolio,
    build_unit_economics,
    normal_cdf,
    rider_payout_at_density,
    threshold_effect,
)


@pytest.fixture(scope="module")
def d():
    return drivers_module.load()


@pytest.fixture(scope="module")
def base(d):
    return build_unit_economics(d)


@pytest.fixture(scope="module")
def decisions(d, base):
    return {dec.key: dec for dec in evaluate_all(d, base)}


# --------------------------------------------------------------------------- #
# 1. Arithmetic identities
# --------------------------------------------------------------------------- #
def test_normal_cdf_matches_known_values():
    assert normal_cdf(0.0) == pytest.approx(0.5)
    assert normal_cdf(1.96) == pytest.approx(0.975, abs=1e-4)
    assert normal_cdf(-1.96) == pytest.approx(0.025, abs=1e-4)


def test_lognormal_is_calibrated_to_the_mean(d):
    """Mu is solved so E[X] equals the AOV driver, not the median."""
    basket = BasketDistribution(d["gross_order_value"], d["basket_log_sigma"])
    assert math.exp(basket.mu + 0.5 * basket.sigma ** 2) == pytest.approx(basket.mean)
    assert basket.median < basket.mean


def test_partial_mean_over_full_support_recovers_the_mean(d):
    """E[X . 1{X < inf}] must integrate back to E[X]."""
    basket = BasketDistribution(d["gross_order_value"], d["basket_log_sigma"])
    assert basket.partial_mean_below(1e9) == pytest.approx(basket.mean, rel=1e-6)


def test_band_mean_lies_inside_its_band(d):
    basket = BasketDistribution(d["gross_order_value"], d["basket_log_sigma"])
    low, high = 374.25, 499.0
    assert low < basket.band_mean(low, high) < high


def test_share_below_is_monotonic(d):
    basket = BasketDistribution(d["gross_order_value"], d["basket_log_sigma"])
    shares = [basket.share_below(t) for t in (100, 200, 300, 400, 500, 600)]
    assert shares == sorted(shares)


def test_ladder_reconciles_to_cm3(base):
    """Each CM level must equal the sum of the lines above it."""
    revenue = (base.net_order_value + base.delivery_fee_income
               + base.handling_fee + base.retail_media_income)
    assert revenue == pytest.approx(base.total_revenue)
    assert base.cm1 == pytest.approx(
        base.total_revenue - base.cogs - base.payment_gateway - base.packaging)
    assert base.cm2 == pytest.approx(
        base.cm1 - base.rider_payout - base.picking_labour - base.spoilage)
    assert base.cm3 == pytest.approx(base.cm2 - base.store_fixed_per_order)


def test_monthly_totals_reconcile_with_per_order_figures(base):
    assert base.cm2_total == pytest.approx(base.cm2 * base.orders_per_month)
    assert base.cm3_total == pytest.approx(base.cm3 * base.orders_per_month, rel=1e-9)
    assert base.ebitda_total == pytest.approx(
        base.ebitda_per_order * base.orders_per_month, rel=1e-9)


def test_retail_media_is_excluded_from_the_payment_gateway_base(d, base):
    """Brand income is invoiced, not collected through the consumer rail."""
    collected = base.net_order_value + base.delivery_fee_income + base.handling_fee
    assert base.payment_gateway == pytest.approx(collected * d["payment_gateway_pct"])
    # Explicitly not charged on the full revenue line.
    assert base.payment_gateway < base.total_revenue * d["payment_gateway_pct"]


def test_cohort_survival_is_non_increasing_and_retention_climbs(d, base):
    portfolio = build_channel_portfolio(d, base.cm3, base.cm2)
    for result in portfolio.results:
        cohort = result.cohort_cm3
        assert cohort.survival[0] == 1.0
        assert cohort.survival == sorted(cohort.survival, reverse=True)
        assert cohort.retention_rate[0] < cohort.retention_rate[-1]
        assert cohort.retention_rate[-1] <= d["asymptotic_retention"]
        assert cohort.cumulative == sorted(cohort.cumulative)


def test_ltv_scales_linearly_with_contribution_margin(d, base):
    """Used by the workbook to derive the CM2 LTV from the CM3 one."""
    portfolio = build_channel_portfolio(d, base.cm3, base.cm2)
    for result in portfolio.results:
        assert result.ltv_cm2 == pytest.approx(
            result.ltv_cm3 * base.cm2 / base.cm3, rel=1e-9)


# --------------------------------------------------------------------------- #
# 2. Modelling conventions
# --------------------------------------------------------------------------- #
def test_fixed_costs_stay_fixed_in_rupees_when_volume_changes(d, base):
    """The convention the store-expansion verdict depends on.

    Halve the volume and the *total* store fixed, central overhead and marketing must
    not move; only the per-order allocations may. If someone replaces this with a
    constant per-order rate, expansion stops looking expensive and this test fails.
    """
    half = build_unit_economics(
        d,
        orders_per_store_per_day=base.orders_per_store_per_day / 2,
        baseline_orders_per_month=base.orders_per_month,
    )
    assert half.store_fixed_total == pytest.approx(base.store_fixed_total)
    assert half.central_overhead_total == pytest.approx(base.central_overhead_total)
    assert half.marketing_total == pytest.approx(base.marketing_total)

    assert half.store_fixed_per_order == pytest.approx(base.store_fixed_per_order * 2)
    assert half.central_overhead_per_order == pytest.approx(
        base.central_overhead_per_order * 2)
    assert half.cm2 == pytest.approx(base.cm2)  # CM2 is volume-independent
    assert half.cm3 < base.cm3


def test_rider_payout_falls_with_store_density(d):
    """Densification is a real benefit and the model must credit it."""
    base_stores = d["dark_stores"]
    assert rider_payout_at_density(d, base_stores) == pytest.approx(d["rider_payout"])
    denser = rider_payout_at_density(d, base_stores + d["new_stores"])
    assert denser < d["rider_payout"]
    # Only the distance-linked share can fall.
    floor = d["rider_payout"] * (1 - d["rider_payout_distance_share"])
    assert denser > floor


def test_up_sizing_raises_basket_and_lowers_the_fee_paying_share(d):
    threshold = d["proposed_threshold"]
    with_upsize = threshold_effect(d, threshold)
    without = threshold_effect(d, threshold, apply_upsizing=False)
    assert with_upsize.gross_order_value > without.gross_order_value
    assert with_upsize.fee_paying_share < without.fee_paying_share
    assert without.gross_order_value == pytest.approx(d["gross_order_value"])


def test_raising_the_threshold_increases_fee_income(d):
    low = threshold_effect(d, d["free_delivery_threshold"])
    high = threshold_effect(d, d["proposed_threshold"])
    assert high.delivery_fee_income > low.delivery_fee_income
    assert high.fee_paying_share > low.fee_paying_share


# --------------------------------------------------------------------------- #
# 3. The headline findings, exactly as the prose states them
# --------------------------------------------------------------------------- #
def test_base_economics_match_the_documented_figures(base):
    assert base.orders_per_month == pytest.approx(604_800)
    assert base.net_order_value == pytest.approx(448.97, abs=0.01)
    assert base.cm1 == pytest.approx(115.62, abs=0.01)
    assert base.cm2 == pytest.approx(63.48, abs=0.01)
    assert base.cm3 == pytest.approx(26.97, abs=0.01)
    assert base.ebitda_per_order == pytest.approx(-6.03, abs=0.01)
    assert base.cm3_total == pytest.approx(16_313_082, abs=1)
    assert base.ebitda_total == pytest.approx(-3_645_318, abs=1)


def test_margins_sit_inside_the_public_benchmark_range(base):
    """CM1 near Blinkit's 26.6% gross margin, CM2 near the ~13% dark-store benchmark."""
    assert 0.24 <= base.pct_of_nov(base.cm1) <= 0.28
    assert 0.12 <= base.pct_of_nov(base.cm2) <= 0.16
    # Contribution-positive but EBITDA-negative: the actual industry position.
    assert base.cm3 > 0
    assert base.ebitda_per_order < 0


def test_store_expansion_is_rejected_and_misses_its_break_even(decisions):
    dec = decisions["store_expansion"]
    assert dec.verdict == "REJECT"
    assert not dec.approved
    assert dec.monthly_delta < 0
    assert dec.payback_months is None
    assert dec.break_even_value == pytest.approx(0.637, abs=0.005)
    assert dec.observed_value == pytest.approx(0.43)
    assert not dec.clears_break_even
    assert dec.margin_of_safety_pp == pytest.approx(-20.7, abs=0.2)


def test_expansion_grows_revenue_while_shrinking_contribution(decisions):
    """The finding. Judged on revenue this proposal is approved; it should not be."""
    dec = decisions["store_expansion"]
    assert dec.order_growth > 0.05
    assert dec.revenue_growth > 0.05
    assert dec.monthly_delta < 0
    assert dec.proposed.cm3 < dec.base.cm3
    assert dec.proposed.store_fixed_per_order > dec.base.store_fixed_per_order
    assert dec.proposed.orders_per_store_per_day < dec.base.orders_per_store_per_day
    # And it is still rejected despite being credited the densification saving.
    assert dec.proposed.rider_payout < dec.base.rider_payout


def test_threshold_change_shrinks_revenue_while_growing_contribution(decisions):
    """The mirror image, which is why revenue cannot be the decision criterion."""
    dec = decisions["threshold_change"]
    assert dec.verdict == "APPROVE"
    assert dec.revenue_growth < 0
    assert dec.order_growth < 0
    assert dec.monthly_delta > 0
    assert dec.proposed.cm3 > dec.base.cm3


def test_threshold_break_even_direction_is_the_right_way_round(decisions):
    """A *smaller* observed decline is favourable here, unlike incrementality."""
    dec = decisions["threshold_change"]
    assert dec.higher_is_better is False
    assert dec.break_even_value == pytest.approx(0.113, abs=0.005)
    assert dec.observed_value == pytest.approx(0.031)
    assert dec.clears_break_even
    assert dec.margin_of_safety_pp == pytest.approx(8.2, abs=0.2)


def test_retail_media_is_the_best_proposal_per_unit_of_risk(decisions):
    dec = decisions["retail_media"]
    assert dec.verdict == "APPROVE"
    assert dec.monthly_delta > decisions["threshold_change"].monthly_delta
    assert dec.order_growth == pytest.approx(0.0)
    # Only a small share of the uplift needs to land.
    assert dec.break_even_value < 0.25


def test_the_two_approved_proposals_are_not_additive(d, base, decisions):
    approved = [dec for dec in decisions.values() if dec.approved]
    assert len(approved) == 2
    combined = evaluate_combined(d, base, approved)
    assert combined.interaction < 0
    assert combined.monthly_delta < combined.sum_of_parts_monthly
    # Adding the business cases overstates the plan by a real but modest amount.
    assert abs(combined.interaction) / combined.sum_of_parts_monthly < 0.05


def test_both_approved_proposals_together_turn_ebitda_positive(d, base, decisions):
    approved = [dec for dec in decisions.values() if dec.approved]
    combined = evaluate_combined(d, base, approved)
    assert base.ebitda_total < 0
    assert combined.combined.ebitda_total > 0


def test_blended_cac_flatters_paid_performance(d, base):
    portfolio = build_channel_portfolio(d, base.cm3, base.cm2)
    assert portfolio.blended_cac < portfolio.paid_cac
    assert portfolio.blended_ratio == pytest.approx(2.20, abs=0.02)
    assert portfolio.paid_ratio == pytest.approx(1.28, abs=0.02)
    overstatement = portfolio.blended_ratio / portfolio.paid_ratio - 1
    assert overstatement == pytest.approx(0.72, abs=0.02)
    assert portfolio.organic_share_of_customers == pytest.approx(0.42, abs=0.01)


def test_two_channels_destroy_value_and_hold_a_large_share_of_spend(d, base):
    portfolio = build_channel_portfolio(d, base.cm3, base.cm2)
    names = {r.name for r in portfolio.sub_unity}
    assert names == {"affiliate_coupon", "offline_ooh"}
    assert portfolio.sub_unity_spend_share == pytest.approx(0.46, abs=0.01)
    for result in portfolio.sub_unity:
        assert result.payback_month is None, f"{result.name} should never pay back"


def test_channel_verdicts_flip_between_cm2_and_cm3(d, base):
    """The reason METRIC_DEFINITIONS.md exists."""
    portfolio = build_channel_portfolio(d, base.cm3, base.cm2)
    flips = {r.name for r in portfolio.results if r.verdict_flips_on_definition}
    assert flips == {"affiliate_coupon", "offline_ooh"}
    for result in portfolio.results:
        if result.name in flips:
            assert result.ratio_cm2 >= 1.0
            assert result.ratio_cm3 < 1.0


def test_organic_has_no_cac_rather_than_a_cac_of_zero(d, base):
    portfolio = build_channel_portfolio(d, base.cm3, base.cm2)
    organic = next(r for r in portfolio.results if r.name == "organic")
    assert organic.cac == 0.0
    assert organic.ratio_cm3 == math.inf
    assert not organic.channel.is_paid


def test_break_even_density_and_the_strategic_reading(d, base):
    be = compute_break_even(d, base)
    assert be.cm3_orders_per_store_per_day == pytest.approx(242, abs=2)
    assert be.ebitda_orders_per_store_per_day == pytest.approx(460, abs=2)
    # CM3 breaks even well below current density; EBITDA does not.
    assert be.cm3_orders_per_store_per_day < base.orders_per_store_per_day
    assert be.ebitda_orders_per_store_per_day > base.orders_per_store_per_day
    # Break-even needs more stores at *current* density, which infill cannot deliver.
    assert be.ebitda_stores_at_current_density > d["dark_stores"]
    assert be.ebitda_stores_at_current_density == pytest.approx(58.7, abs=0.5)


def test_scenarios_are_ordered_and_bear_is_loss_making(d, base):
    scenarios = {s.label: s for s in build_scenarios(d, base)}
    assert set(scenarios) == {"Base", "Bull", "Bear"}
    assert scenarios["Bear"].cm3 < scenarios["Base"].cm3 < scenarios["Bull"].cm3
    assert scenarios["Bull"].ebitda_total > 0 > scenarios["Bear"].ebitda_total


def test_sensitivity_grid_is_anchored_and_monotonic(d, base):
    grid = cm3_sensitivity(d, base)
    mid_row = grid.row_values.index(420.0)
    mid_col = grid.col_values.index(0.215)
    assert grid.cells[mid_row][mid_col] == pytest.approx(base.cm3, abs=0.01)

    for row in grid.cells:
        assert row == sorted(row), "CM3 must rise with gross margin"
    for column in range(len(grid.col_values)):
        assert [r[column] for r in grid.cells] == sorted(r[column] for r in grid.cells), (
            "CM3 must rise with order density"
        )


def test_margin_wins_on_cm3_and_density_wins_on_ebitda(d, base):
    """The ranking of the two big levers reverses between CM3 and EBITDA.

    Density does double duty: it lifts CM3 and dilutes fixed central overhead across
    more orders. Margin only does the first. An earlier draft of the analysis claimed
    density beat margin outright, which is false on CM3 -- hence this test.
    """
    reference = build_unit_economics(d, baseline_orders_per_month=base.orders_per_month)
    plus_margin = build_unit_economics(
        d, product_gross_margin_pct=d["product_gross_margin_pct"] + 0.01,
        baseline_orders_per_month=base.orders_per_month,
    )
    plus_density = build_unit_economics(
        d, orders_per_store_per_day=base.orders_per_store_per_day + 40,
        baseline_orders_per_month=base.orders_per_month,
    )

    margin_cm3 = plus_margin.cm3 - reference.cm3
    density_cm3 = plus_density.cm3 - reference.cm3
    margin_ebitda = plus_margin.ebitda_per_order - reference.ebitda_per_order
    density_ebitda = plus_density.ebitda_per_order - reference.ebitda_per_order

    assert margin_cm3 == pytest.approx(4.49, abs=0.02)
    assert density_cm3 == pytest.approx(3.17, abs=0.02)
    assert margin_ebitda == pytest.approx(4.49, abs=0.02)
    assert density_ebitda == pytest.approx(6.04, abs=0.02)

    assert margin_cm3 > density_cm3, "margin should win on CM3"
    assert density_ebitda > margin_ebitda, "density should win on EBITDA"
    # Margin passes straight through: it does nothing to overhead absorption.
    assert margin_ebitda == pytest.approx(margin_cm3, abs=1e-9)
    # Density does not: the gain is larger on EBITDA than on CM3.
    assert density_ebitda > density_cm3


def test_expansion_grid_is_negative_everywhere_at_measured_incrementality(d, base):
    grid = expansion_sensitivity(d, base)
    row = grid.row_values.index(0.40)
    assert all(cell < 0 for cell in grid.cells[row]), (
        "at 40% incrementality -- close to the pilot's measured 43% -- every "
        "new-store volume assumption should destroy contribution"
    )
    # Higher incrementality is monotonically better.
    for column in range(len(grid.col_values)):
        assert [r[column] for r in grid.cells] == sorted(r[column] for r in grid.cells)


# --------------------------------------------------------------------------- #
# 4. Driver validation
# --------------------------------------------------------------------------- #
def test_every_driver_states_a_basis(d):
    assert all(driver.basis for driver in d.all())
    assert len(d.assumptions) >= 3
    assert {dr.name for dr in d.measured} == {"incrementality", "observed_volume_response"}


def mutate(tmp_path, edit):
    """Write a mutated copy of the driver file and load it."""
    raw = yaml.safe_load(drivers_module.CONFIG_PATH.read_text(encoding="utf-8"))
    edit(raw)
    path = tmp_path / "drivers.yml"
    path.write_text(yaml.safe_dump(raw), encoding="utf-8")
    return drivers_module.load(path)


def test_a_driver_without_a_basis_is_rejected(tmp_path):
    with pytest.raises(ValueError, match="no stated basis"):
        mutate(tmp_path, lambda raw: raw["basket"]["gross_order_value"].update(basis=""))


def test_a_fraction_outside_zero_to_one_is_rejected(tmp_path):
    with pytest.raises(ValueError, match="fraction but reads"):
        mutate(tmp_path, lambda raw: raw["revenue"]["product_gross_margin_pct"].update(value=1.4))


def test_a_non_positive_volume_driver_is_rejected(tmp_path):
    with pytest.raises(ValueError, match="must be positive"):
        mutate(tmp_path, lambda raw: raw["scale"]["orders_per_store_per_day"].update(value=0))


def test_an_inverted_retention_curve_is_rejected(tmp_path):
    with pytest.raises(ValueError, match="below asymptotic_retention"):
        mutate(tmp_path, lambda raw: raw["cohort"]["month_1_retention"].update(value=0.95))


def test_a_threshold_proposal_that_is_not_an_increase_is_rejected(tmp_path):
    with pytest.raises(ValueError, match="not above the current threshold"):
        mutate(
            tmp_path,
            lambda raw: raw["decisions"]["threshold_change"]["proposed_threshold"].update(
                value=250.0),
        )


def test_an_unknown_driver_name_raises(d):
    with pytest.raises(KeyError, match="unknown driver"):
        _ = d["not_a_driver"]


# --------------------------------------------------------------------------- #
# 5. The delivered workbook
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def evaluated_workbook(d, tmp_path_factory):
    """Build the xlsx and evaluate it with an independent formula engine.

    This is the test that matters most. Everything above verifies the Python. This
    verifies the artefact a recruiter actually opens, using a different engine from
    the one that wrote it.
    """
    formulas = pytest.importorskip("formulas", reason="needs `formulas` to evaluate the xlsx")
    from src.workbook import build

    path = tmp_path_factory.mktemp("wb") / "unit_economics.xlsx"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        build(d, path)
        model = formulas.ExcelModel().loads(str(path)).finish()
        solution = model.calculate()

    values: dict[str, float] = {}
    errors: dict[str, str] = {}
    stem = path.name.upper().replace(".XLSX", "")
    for key, cell in solution.items():
        if stem not in key.upper() or ":" in key:
            continue
        try:
            value = cell.value[0, 0]
        except Exception:
            # Ranges and non-scalar results are not of interest here.
            continue
        text = str(value)
        if text.startswith("#"):
            errors[key] = text
        name = key.split("!")[-1]
        values[name.upper()] = value
    return values, errors


def named(values, key):
    """Named ranges resolve to the cell they point at, so look them up by name."""
    for candidate in (key.upper(), f"${key.upper()}"):
        if candidate in values:
            return values[candidate]
    matches = [v for k, v in values.items() if k == key.upper()]
    assert matches, f"named range {key!r} not found in the evaluated workbook"
    return matches[0]


def test_workbook_has_no_error_cells(evaluated_workbook):
    _, errors = evaluated_workbook
    assert errors == {}, f"workbook contains error cells: {errors}"


def test_workbook_basket_sheet_matches_python(d, evaluated_workbook):
    values, _ = evaluated_workbook
    current = threshold_effect(d, d["free_delivery_threshold"])
    proposed = threshold_effect(d, d["proposed_threshold"])
    assert named(values, "gov_current") == pytest.approx(current.gross_order_value, rel=1e-9)
    assert named(values, "feeinc_current") == pytest.approx(current.delivery_fee_income, rel=1e-9)
    assert named(values, "feeshare_current") == pytest.approx(current.fee_paying_share, rel=1e-9)
    assert named(values, "gov_proposed") == pytest.approx(proposed.gross_order_value, rel=1e-9)
    assert named(values, "feeinc_proposed") == pytest.approx(
        proposed.delivery_fee_income, rel=1e-9)


def test_workbook_ladder_matches_python(base, evaluated_workbook):
    values, _ = evaluated_workbook
    assert named(values, "base_nov") == pytest.approx(base.net_order_value, rel=1e-9)
    assert named(values, "base_cm1") == pytest.approx(base.cm1, rel=1e-9)
    assert named(values, "base_cm2") == pytest.approx(base.cm2, rel=1e-9)
    assert named(values, "base_cm3") == pytest.approx(base.cm3, rel=1e-9)
    assert named(values, "base_orders") == pytest.approx(base.orders_per_month, rel=1e-9)
    assert named(values, "base_cm2_total") == pytest.approx(base.cm2_total, rel=1e-9)
    assert named(values, "base_cm3_total") == pytest.approx(base.cm3_total, rel=1e-9)
    assert named(values, "base_ebitda_total") == pytest.approx(base.ebitda_total, rel=1e-9)
    assert named(values, "base_ebitda_per_order") == pytest.approx(
        base.ebitda_per_order, rel=1e-9)


def test_workbook_decision_columns_match_python(decisions, evaluated_workbook):
    values, _ = evaluated_workbook
    assert named(values, "exp_cm3_total") == pytest.approx(
        decisions["store_expansion"].proposed.cm3_total, rel=1e-9)
    assert named(values, "thr_cm3_total") == pytest.approx(
        decisions["threshold_change"].proposed.cm3_total, rel=1e-9)
    assert named(values, "media_cm3_total") == pytest.approx(
        decisions["retail_media"].proposed.cm3_total, rel=1e-9)


def test_workbook_reproduces_the_verdicts(base, evaluated_workbook):
    """The signs, not just the magnitudes -- this is what a reader acts on."""
    values, _ = evaluated_workbook
    base_total = named(values, "base_cm3_total")
    assert named(values, "exp_cm3_total") - base_total < 0
    assert named(values, "thr_cm3_total") - base_total > 0
    assert named(values, "media_cm3_total") - base_total > 0
    assert base_total == pytest.approx(base.cm3_total, rel=1e-9)


def test_workbook_channel_ltv_and_ratios_match_python(d, base, evaluated_workbook):
    values, _ = evaluated_workbook
    portfolio = build_channel_portfolio(d, base.cm3, base.cm2)
    for result in portfolio.results:
        assert named(values, f"ltv_{result.name}") == pytest.approx(result.ltv_cm3, rel=1e-9)
    assert named(values, "blended_ltv_cac") == pytest.approx(portfolio.blended_ratio, rel=1e-9)
    assert named(values, "paid_ltv_cac") == pytest.approx(portfolio.paid_ratio, rel=1e-9)


def test_workbook_has_every_expected_sheet(d, tmp_path):
    from openpyxl import load_workbook

    from src.workbook import build

    path = build(d, tmp_path / "wb.xlsx")
    sheets = load_workbook(path).sheetnames
    assert sheets == [
        "Guide", "Drivers", "Basket", "UnitEconomics",
        "Cohort", "Channels", "Decisions", "Scenarios", "Sensitivity",
    ]
    # Sheet names must not contain spaces: quoting them in formulas is a live
    # failure mode that produces a workbook which opens but does not calculate.
    assert all(" " not in sheet for sheet in sheets)


def test_calculation_sheets_contain_no_hard_coded_numbers(d, tmp_path):
    """Every value cell outside Drivers must be a formula.

    A constant buried in a calculation sheet means the Drivers sheet is no longer the
    whole input surface, and a reviewer can no longer trust any of it.
    """
    from openpyxl import load_workbook

    from src.workbook import build

    path = build(d, tmp_path / "wb.xlsx")
    workbook = load_workbook(path)
    # Sensitivity, Scenarios and Cohort carry grid axis labels (densities, margins,
    # month numbers), which are inputs to the grid by design rather than results.
    checked = ("Basket", "UnitEconomics", "Channels", "Decisions")
    offenders = [
        f"{name}!{cell.coordinate} = {cell.value}"
        for name in checked
        for row in workbook[name].iter_rows(min_col=2)
        for cell in row
        if isinstance(cell.value, int | float)
    ]
    assert offenders == [], f"hard-coded values outside Drivers: {offenders}"
